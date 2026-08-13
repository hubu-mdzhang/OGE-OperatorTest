from __future__ import annotations

import csv
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from src.models import OperatorCase
from src.oge_page import OGEDevelopPage
from src.runner import run_cases
from tests.mock_oge import (
    HTML,
    find_browser_executable,
    install_mock_routes,
    playwright_chromium_ready,
)


def config_path(tmp_path: Path, browser_exe: str | None = None) -> Path:
    cfg = {
        "app": {
            "url": "https://mock.oge/develop",
            "browser_channel": None,
            "browser_executable_path": browser_exe,
            "fallback_to_playwright_chromium": False,
            "headless": True,
            "viewport": {"width": 1200, "height": 900},
            "persistent_profile_dir": "runtime/profile",
            "navigation_timeout_ms": 15000,
            "action_timeout_ms": 10000,
            "workspace_wait_ms": 10000,
            "operator_timeout_sec": 8,
            "network_settle_timeout_sec": 4,
            "pre_case_settle_ms": 20,
            "render_settle_ms": 50,
            "reload_before_each_case": True,
            "max_attempts": 1,
            "retry_final_statuses": ["FAIL", "TIMEOUT"],
        },
        "auth": {"login_url_markers": ["/login"], "http_status_codes": [401, 403]},
        "selectors": {
            "editor_root": '.monaco-editor[role="code"]',
            "editor_textarea": '.monaco-editor textarea.inputarea',
            "run_button_primary": 'div[class*="controlButton_oge_editor_control_btn__"]:has(img[src="/svgs/run.svg"])',
            "run_button_fallback": 'img[src="/svgs/run.svg"]',
            "console_root": 'div[class*="console_oge_console__"]',
            "console_entry": 'div[class*="console_entry__"]',
            "console_content": 'span[class*="console_content__"]',
            "globe_canvas": 'canvas[style*="image-rendering: pixelated"]',
        },
        "network": {
            "execute_code_path": "/api/computation-api/executeCode",
            "execute_dag_path": "/api/computation-api/executeDag",
        },
        "judging": {
            "running_text_contains": "正在执行python代码模版",
            "success_text_contains": "运行成功",
            "require_running_before_success": True,
            "screenshot_diff_threshold": 0.01,
        },
        "output": {"root_dir": "output", "report_name": "results.xlsx", "save_trace": True},
    }
    path = tmp_path / "config.yaml"
    path.write_text(yaml.safe_dump(cfg, allow_unicode=True, sort_keys=False), encoding="utf-8")
    return path


def input_csv(tmp_path: Path, cases: list[OperatorCase]) -> Path:
    path = tmp_path / "operators.csv"
    fields = [
        "case_id",
        "category",
        "name_cn",
        "operator_name",
        "code",
        "expected_result_type",
        "enabled",
        "source_status",
        "validation_mode",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for case in cases:
            writer.writerow({field: getattr(case, field) for field in fields})
    return path


def test_full_runner_against_mock_oge(tmp_path, monkeypatch):
    browser_exe = find_browser_executable()
    with sync_playwright() as playwright:
        if not browser_exe and not playwright_chromium_ready(playwright):
            pytest.skip("当前环境没有可用 Chromium/Edge")

    def mock_open(self: OGEDevelopPage, timeout_ms=None):
        install_mock_routes(self.page)
        self.page.set_content(HTML, wait_until="load")
        return self.assert_workspace()

    def mock_reload(self: OGEDevelopPage):
        self.page.set_content(HTML, wait_until="load")
        return self.assert_workspace()

    monkeypatch.setattr(OGEDevelopPage, "open", mock_open)
    monkeypatch.setattr(OGEDevelopPage, "reload_isolated", mock_reload)

    cfg = config_path(tmp_path, browser_exe)
    cases = [
        OperatorCase(
            1,
            "mock",
            "地图成功待复核",
            "Coverage.mockMap",
            "import oge\n# ERROR_WORD_ONLY\n# SUCCESS_BEFORE_DAGS\nprint('x')",
            expected_result_type="MAP",
        ),
        OperatorCase(
            2,
            "mock",
            "日志执行即通过",
            "Coverage.mockLog",
            "import oge\n# LOG_ONLY\nprint('log')",
            expected_result_type="LOG",
            validation_mode="EXECUTION_ONLY",
        ),
        OperatorCase(
            3,
            "mock",
            "DAG失败",
            "Coverage.mockFail",
            "import oge\n# DAG_FAIL\nprint('fail')",
            expected_result_type="MAP",
        ),
    ]
    csv_path = input_csv(tmp_path, cases)
    results = run_cases(cases, cfg, csv_path)

    assert [result.final_status for result in results] == ["REVIEW", "PASS", "FAIL"]
    first = results[0]
    assert first.execution_status == "SUCCESS"
    assert first.result_status == "UNCERTAIN"
    assert first.code_readback_match is True
    assert first.code_payload_match is True
    assert first.network_complete is True
    assert first.dag_count == 3 and first.dag_success_count == 3 and first.dag_missing_count == 0
    assert any("Spatial Error Model" in entry.text for entry in first.console_entries)

    third = results[2]
    assert third.execution_status == "FAIL"
    assert third.dag_failed_count == 1

    run_root = Path(first.evidence_dir)
    if not run_root.is_absolute():
        run_root = Path(cfg).parent / "output" / "runs" / first.run_id / first.evidence_dir
    run_root = run_root.parents[1]
    assert (run_root / "results.jsonl").exists()
    assert (run_root / "results.csv").exists()
    assert len((run_root / "results.jsonl").read_text(encoding="utf-8").splitlines()) == 3
    report = run_root / "results.xlsx"
    assert report.exists()
    workbook = load_workbook(report, data_only=False)
    assert workbook["测试结果"].max_row == 4
    assert workbook["复核队列"].max_row == 2

    for result in results:
        evidence = run_root / result.evidence_dir
        for filename in (
            "source.py",
            "result.json",
            "console.txt",
            "network.json",
            "trace.zip",
            "result_screenshot.png",
            "globe_result.png",
        ):
            assert (evidence / filename).exists(), (result.case_id, filename)


def test_skips_still_get_complete_evidence_and_full_total(tmp_path):
    cfg = config_path(tmp_path)
    cases = [
        OperatorCase(1, "mock", "缺代码", "", "", source_status="MISSING_OPERATOR_NAME_AND_CODE"),
        OperatorCase(2, "mock", "被筛选", "Coverage.ready", "print('x')"),
    ]
    csv_path = input_csv(tmp_path, cases)
    results = run_cases(cases, cfg, csv_path, start_id=99, end_id=100)
    assert [result.final_status for result in results] == ["SKIPPED_NO_CODE", "SKIPPED_FILTERED"]
    run_root = Path(cfg).parent / "output" / "runs" / results[0].run_id
    assert len((run_root / "results.jsonl").read_text(encoding="utf-8").splitlines()) == 2
    workbook = load_workbook(run_root / "results.xlsx", data_only=False)
    assert workbook["测试结果"].max_row == 3
    for result in results:
        evidence = run_root / result.evidence_dir
        assert (evidence / "trace.zip").stat().st_size > 0
        assert (evidence / "result_screenshot.png").stat().st_size > 0
        assert (evidence / "globe_result.png").stat().st_size > 0
