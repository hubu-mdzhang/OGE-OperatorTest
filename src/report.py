from __future__ import annotations

import csv
import json
import os
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

from .models import CaseResult, OperatorCase


CSV_FIELDS = [
    "event_seq",
    "run_id",
    "case_id",
    "category",
    "name_cn",
    "operator_name",
    "source_original_status",
    "source_development_status",
    "source_manual_test_status",
    "source_status",
    "enabled",
    "expected_result_type",
    "attempt",
    "retry_count",
    "execution_status",
    "result_status",
    "final_status",
    "started_at",
    "finished_at",
    "duration_sec",
    "batch_status",
    "execute_code_seen",
    "execute_code_status",
    "network_complete",
    "dag_count",
    "dag_success_count",
    "dag_failed_count",
    "dag_missing_count",
    "code_injection_mode",
    "code_readback_match",
    "code_payload_match",
    "visual_change_ratio",
    "auth_signal",
    "failure_reason",
    "evidence_dir",
    "source_path",
    "result_path",
    "result_screenshot_path",
    "globe_result_path",
    "console_path",
    "network_path",
    "trace_path",
    "console_text",
    "network_summary_json",
]


RESULT_HEADERS = [
    "算子编号",
    "中文名",
    "英文名",
    "分类",
    "原表状态",
    "开发状态",
    "原人工测试状态",
    "输入源状态",
    "启用",
    "预期结果类型",
    "本轮执行状态",
    "结果状态",
    "最终状态",
    "开始时间",
    "测试时间",
    "耗时(s)",
    "重试次数",
    "DAG数量",
    "DAG成功数",
    "DAG失败数",
    "DAG缺失数",
    "代码回读一致",
    "提交代码一致",
    "失败/复核原因",
    "全页面截图",
    "Globe截图",
    "源代码",
    "Result JSON",
    "Console",
    "Network",
    "Trace",
    "证据目录",
]


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def atomic_write_json(path: str | Path, payload: Any) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp = target.with_suffix(target.suffix + ".tmp")
    temp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, target)


def save_case_json(result: CaseResult, path: str | Path) -> None:
    atomic_write_json(path, result.to_dict())


def _flat_result(result: CaseResult, event_seq: int) -> dict[str, Any]:
    return {
        "event_seq": event_seq,
        "run_id": result.run_id,
        "case_id": result.case_id,
        "category": result.category,
        "name_cn": result.name_cn,
        "operator_name": result.operator_name,
        "source_original_status": result.source_original_status,
        "source_development_status": result.source_development_status,
        "source_manual_test_status": result.source_manual_test_status,
        "source_status": result.source_status,
        "enabled": result.enabled,
        "expected_result_type": result.expected_result_type,
        "attempt": result.attempt,
        "retry_count": result.retry_count,
        "execution_status": result.execution_status,
        "result_status": result.result_status,
        "final_status": result.final_status,
        "started_at": result.started_at,
        "finished_at": result.finished_at,
        "duration_sec": round(result.duration_sec, 3),
        "batch_status": result.batch_status,
        "execute_code_seen": result.execute_code_seen,
        "execute_code_status": result.execute_code_status,
        "network_complete": result.network_complete,
        "dag_count": result.dag_count,
        "dag_success_count": result.dag_success_count,
        "dag_failed_count": result.dag_failed_count,
        "dag_missing_count": result.dag_missing_count,
        "code_injection_mode": result.code_injection_mode,
        "code_readback_match": result.code_readback_match,
        "code_payload_match": result.code_payload_match,
        "visual_change_ratio": result.visual_change_ratio,
        "auth_signal": result.auth_signal,
        "failure_reason": result.failure_reason,
        "evidence_dir": result.evidence_dir,
        "source_path": result.source_path,
        "result_path": result.result_path,
        "result_screenshot_path": result.result_screenshot_path,
        "globe_result_path": result.globe_result_path,
        "console_path": result.console_path,
        "network_path": result.network_path,
        "trace_path": result.trace_path,
        "console_text": "\n".join(f"[{entry.type}] {entry.text}" for entry in result.console_entries),
        "network_summary_json": json.dumps(result.network_summary, ensure_ascii=False, separators=(",", ":")),
    }


def append_result_event(result: CaseResult, jsonl_path: str | Path, csv_path: str | Path) -> int:
    jsonl = Path(jsonl_path)
    csv_file = Path(csv_path)
    jsonl.parent.mkdir(parents=True, exist_ok=True)
    existing_events = read_jsonl_events(jsonl)
    event_seq = len(existing_events) + 1
    payload = result.to_dict()
    payload["event_seq"] = event_seq

    with jsonl.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")
        handle.flush()
        os.fsync(handle.fileno())

    write_header = not csv_file.exists() or csv_file.stat().st_size == 0
    with csv_file.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS, extrasaction="ignore")
        if write_header:
            writer.writeheader()
        writer.writerow(_flat_result(result, event_seq))
        handle.flush()
        os.fsync(handle.fileno())
    return event_seq


def read_jsonl_events(path: str | Path) -> list[dict[str, Any]]:
    source = Path(path)
    if not source.exists():
        return []
    events: list[dict[str, Any]] = []
    for line_number, line in enumerate(source.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError:
            # A process can be killed between write and fsync. Ignore only the final torn line.
            if line_number == len(source.read_text(encoding="utf-8").splitlines()):
                break
            raise
        if isinstance(payload, dict):
            events.append(payload)
    return events


def latest_results(path: str | Path) -> dict[int, dict[str, Any]]:
    latest: dict[int, dict[str, Any]] = {}
    for event in read_jsonl_events(path):
        try:
            latest[int(event["case_id"])] = event
        except (KeyError, TypeError, ValueError):
            continue
    return latest


def write_console(entries: Iterable[Any], path: str | Path) -> None:
    lines = [f"[{entry.index}] [{entry.type}] {entry.text}" for entry in entries]
    Path(path).write_text("\n".join(lines), encoding="utf-8")


def write_network(records: Iterable[Any], path: str | Path) -> None:
    payload = []
    for record in records:
        payload.append(record if isinstance(record, dict) else record.__dict__)
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _placeholder_png(path: Path, case: OperatorCase, status: str, detail: str) -> None:
    image = Image.new("RGB", (1400, 900), color=(245, 247, 250))
    draw = ImageDraw.Draw(image)
    lines = [
        "OGE OPERATOR AUTOMATION EVIDENCE PLACEHOLDER",
        f"case_id: {case.case_id}",
        f"operator: {case.operator_name or '[missing]'}",
        f"status: {status}",
        f"detail: {detail[:180]}",
        f"generated_at: {now_iso()}",
    ]
    draw.rectangle((60, 60, 1340, 840), outline=(80, 95, 115), width=3)
    for index, line in enumerate(lines):
        draw.text((110, 130 + index * 90), line, fill=(30, 41, 59))
    image.save(path)


def ensure_required_evidence(
    evidence_dir: str | Path,
    case: OperatorCase,
    status: str,
    detail: str,
) -> dict[str, Path]:
    root = Path(evidence_dir)
    root.mkdir(parents=True, exist_ok=True)
    paths = {
        "source": root / "source.py",
        "result": root / "result.json",
        "console": root / "console.txt",
        "network": root / "network.json",
        "trace": root / "trace.zip",
        "result_screenshot": root / "result_screenshot.png",
        "globe_result": root / "globe_result.png",
    }
    if not paths["source"].exists():
        source_text = case.code if case.code else f"# SKIPPED_NO_CODE: case_id={case.case_id}, operator_name unavailable\n"
        paths["source"].write_text(source_text, encoding="utf-8")
    if not paths["console"].exists():
        paths["console"].write_text(f"[{status}] {detail}\n", encoding="utf-8")
    if not paths["network"].exists():
        paths["network"].write_text("[]\n", encoding="utf-8")
    if not paths["trace"].exists():
        with zipfile.ZipFile(paths["trace"], "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("README.txt", f"No Playwright trace was available.\nstatus={status}\ndetail={detail}\n")
    if not paths["result_screenshot"].exists():
        _placeholder_png(paths["result_screenshot"], case, status, detail)
    if not paths["globe_result"].exists():
        _placeholder_png(paths["globe_result"], case, status, detail)
    return paths


def _event_value(event: dict[str, Any] | None, key: str, default: Any = "") -> Any:
    if event is None:
        return default
    value = event.get(key, default)
    return default if value is None else value


def _result_row(case: OperatorCase, event: dict[str, Any] | None) -> list[Any]:
    final_status = _event_value(event, "final_status", "PENDING")
    return [
        case.case_id,
        case.name_cn,
        case.operator_name,
        case.category,
        case.source_original_status,
        case.source_development_status,
        case.source_manual_test_status,
        case.source_status,
        case.enabled,
        case.expected_result_type,
        _event_value(event, "execution_status", "NOT_RUN"),
        _event_value(event, "result_status", "NOT_EVALUATED"),
        final_status,
        _event_value(event, "started_at"),
        _event_value(event, "finished_at"),
        _event_value(event, "duration_sec", 0),
        _event_value(event, "retry_count", 0),
        _event_value(event, "dag_count", 0),
        _event_value(event, "dag_success_count", 0),
        _event_value(event, "dag_failed_count", 0),
        _event_value(event, "dag_missing_count", 0),
        _event_value(event, "code_readback_match"),
        _event_value(event, "code_payload_match"),
        _event_value(event, "failure_reason"),
        _event_value(event, "result_screenshot_path"),
        _event_value(event, "globe_result_path"),
        _event_value(event, "source_path"),
        _event_value(event, "result_path"),
        _event_value(event, "console_path"),
        _event_value(event, "network_path"),
        _event_value(event, "trace_path"),
        _event_value(event, "evidence_dir"),
    ]


def generate_results_xlsx(
    cases: list[OperatorCase],
    jsonl_path: str | Path,
    report_path: str | Path,
    run_metadata: dict[str, Any],
) -> Path:
    report = Path(report_path)
    report.parent.mkdir(parents=True, exist_ok=True)
    latest = latest_results(jsonl_path)
    final_counts = Counter(str(event.get("final_status") or "PENDING") for event in latest.values())
    pending_count = len(cases) - len(latest)
    if pending_count:
        final_counts["PENDING"] += pending_count

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "OGE 算子批量自动化测试结果"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary.merge_cells("A1:D1")
    summary["A3"] = "Run ID"
    summary["B3"] = run_metadata.get("run_id", "")
    summary["A4"] = "批次状态"
    summary["B4"] = run_metadata.get("batch_status", "")
    summary["A5"] = "输入 CSV"
    summary["B5"] = run_metadata.get("input_csv", "")
    summary["A6"] = "更新时间"
    summary["B6"] = now_iso()

    summary["A8"] = "业务口径"
    summary["B8"] = "数量"
    summary["A9"] = "总记录数"
    summary["B9"] = len(cases)
    summary["A10"] = "可执行记录数"
    summary["B10"] = sum(1 for case in cases if case.enabled and case.has_code)
    summary["A11"] = "缺代码记录数"
    summary["B11"] = sum(1 for case in cases if not case.has_code)

    statuses = [
        "PASS",
        "FAIL",
        "REVIEW",
        "TIMEOUT",
        "SKIPPED_NO_CODE",
        "SKIPPED_DISABLED",
        "SKIPPED_FILTERED",
        "SKIPPED_AUTH_EXPIRED",
        "SKIPPED_BATCH_ABORTED",
        "PENDING",
    ]
    summary["A13"] = "最终状态"
    summary["B13"] = "数量"
    for offset, status in enumerate(statuses, start=14):
        summary.cell(offset, 1, status)
        summary.cell(offset, 2, final_counts.get(status, 0))
    reconciliation_row = 14 + len(statuses)
    summary.cell(reconciliation_row, 1, "总数 - 全部最终状态")
    summary.cell(reconciliation_row, 2, f"=B9-SUM(B14:B{reconciliation_row - 1})")
    summary.cell(reconciliation_row, 3, "应为 0")
    for row in (8, 13):
        summary[f"A{row}:B{row}"][0][0].fill = PatternFill("solid", fgColor="D9EAF7")
        summary[f"A{row}:B{row}"][0][1].fill = PatternFill("solid", fgColor="D9EAF7")
        summary[f"A{row}:B{row}"][0][0].font = Font(bold=True)
        summary[f"A{row}:B{row}"][0][1].font = Font(bold=True)
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 70
    summary.column_dimensions["C"].width = 16

    results = workbook.create_sheet("测试结果")
    results.sheet_view.showGridLines = False
    results.append(RESULT_HEADERS)
    for case in cases:
        results.append(_result_row(case, latest.get(case.case_id)))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in results[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    results.freeze_panes = "D2"
    results.auto_filter.ref = f"A1:{get_column_letter(len(RESULT_HEADERS))}{results.max_row}"
    widths = [10, 24, 48, 18, 32, 18, 24, 30, 9, 18, 18, 18, 24, 23, 23, 12, 12, 10, 12, 12, 12, 15, 15, 62, 35, 35, 35, 35, 35, 35, 35, 40]
    for index, width in enumerate(widths, start=1):
        results.column_dimensions[get_column_letter(index)].width = width
    for row in results.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(2, results.max_row + 1):
        for column_number in range(25, 33):
            cell = results.cell(row_number, column_number)
            if cell.value:
                cell.hyperlink = str(cell.value).replace("\\", "/")
                cell.style = "Hyperlink"

    status_col = "M"
    status_range = f"{status_col}2:{status_col}{results.max_row}"
    status_colors = {
        "PASS": "C6EFCE",
        "FAIL": "FFC7CE",
        "REVIEW": "FFEB9C",
        "TIMEOUT": "F4B183",
        "SKIPPED": "D9E1F2",
    }
    for status, color in status_colors.items():
        formula = f'LEFT(${status_col}2,{len(status)})="{status}"'
        results.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
        )

    review = workbook.create_sheet("复核队列")
    review.sheet_view.showGridLines = False
    review_headers = ["算子编号", "中文名", "英文名", "复核原因", "全页面截图", "Globe截图", "Result JSON"]
    review.append(review_headers)
    for case in cases:
        event = latest.get(case.case_id)
        if not event or event.get("final_status") != "REVIEW":
            continue
        review.append(
            [
                case.case_id,
                case.name_cn,
                case.operator_name,
                event.get("failure_reason", ""),
                event.get("result_screenshot_path", ""),
                event.get("globe_result_path", ""),
                event.get("result_path", ""),
            ]
        )
    for cell in review[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:G{max(1, review.max_row)}"
    for column, width in zip("ABCDEFG", [10, 24, 48, 70, 38, 38, 38]):
        review.column_dimensions[column].width = width
    for row_number in range(2, review.max_row + 1):
        for column_number in range(5, 8):
            cell = review.cell(row_number, column_number)
            if cell.value:
                cell.hyperlink = str(cell.value).replace("\\", "/")
                cell.style = "Hyperlink"
        for cell in review[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    temp = report.with_suffix(report.suffix + ".tmp")
    workbook.save(temp)
    os.replace(temp, report)
    return report
