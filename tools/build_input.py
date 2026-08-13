from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELDNAMES = [
    "case_id",
    "category",
    "name_cn",
    "operator_name",
    "code",
    "expected_result_type",
    "enabled",
    "source_status",
    "source_original_status",
    "source_development_status",
    "source_manual_test_status",
    "description",
    "input_data_type",
    "source_notes",
    "validation_mode",
    "expected_console_regex",
    "source_file",
    "source_sheet",
    "source_row",
]

REQUIRED_HEADERS = [
    "编号",
    "分类",
    "英文名",
    "中文名",
    "算子描述",
    "输入数据类型",
    "状态（原表）",
    "开发状态",
    "测试状态",
    "示例代码",
]


def text(value: Any) -> str:
    return str(value or "").strip()


def infer_result_type(code: str) -> str:
    has_map = any(token in code for token in (".getMap(", ".centerMap(", ".centerObject(", "addLayer("))
    has_log = any(token in code for token in (".log(", "print(", ".print("))
    if has_map and has_log:
        return "MAP_AND_LOG"
    if has_map:
        return "MAP"
    if has_log:
        return "LOG"
    return "UNKNOWN"


def status_for(operator_name: str, code: str) -> str:
    if not operator_name and not code:
        return "MISSING_OPERATOR_NAME_AND_CODE"
    if not operator_name:
        return "MISSING_OPERATOR_NAME"
    if not code:
        return "MISSING_CODE"
    return "READY"


def parse_case_id(value: Any, source_row: int) -> int:
    if value in (None, ""):
        raise ValueError(f"Excel 第 {source_row} 行编号为空")
    try:
        number = int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Excel 第 {source_row} 行编号不是整数: {value!r}") from exc
    if number <= 0:
        raise ValueError(f"Excel 第 {source_row} 行编号必须大于 0: {number}")
    return number


def build_rows(excel_path: Path, sheet_name: str | None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header_values = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_map = {name: index for index, name in enumerate(header_values) if name}
    missing_headers = [name for name in REQUIRED_HEADERS if name not in header_map]
    if missing_headers:
        raise ValueError(f"Excel 缺少必需列: {', '.join(missing_headers)}")

    code_index = header_map["示例代码"]
    rows: list[dict[str, Any]] = []
    syntax_errors: list[dict[str, Any]] = []

    for source_row, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(cells)
        if not any(value not in (None, "") for value in values):
            continue
        case_id = parse_case_id(values[header_map["编号"]], source_row)
        operator_name = text(values[header_map["英文名"]])
        code = str(values[code_index] or "").replace("\r\n", "\n").replace("\r", "\n").rstrip("\n")
        source_notes = text(values[code_index + 1]) if code_index + 1 < len(values) else ""
        source_status = status_for(operator_name, code)

        if code:
            try:
                compile(code, f"<{excel_path.name}:{source_row}>", "exec")
            except SyntaxError as exc:
                syntax_errors.append(
                    {"case_id": case_id, "source_row": source_row, "message": str(exc)}
                )

        rows.append(
            {
                "case_id": case_id,
                "category": text(values[header_map["分类"]]),
                "name_cn": text(values[header_map["中文名"]]),
                "operator_name": operator_name,
                "code": code,
                "expected_result_type": infer_result_type(code) if code else "UNKNOWN",
                "enabled": "true",
                "source_status": source_status,
                "source_original_status": text(values[header_map["状态（原表）"]]),
                "source_development_status": text(values[header_map["开发状态"]]),
                "source_manual_test_status": text(values[header_map["测试状态"]]),
                "description": text(values[header_map["算子描述"]]),
                "input_data_type": text(values[header_map["输入数据类型"]]),
                "source_notes": source_notes,
                "validation_mode": "MANUAL_OR_MULTIMODAL",
                "expected_console_regex": "",
                "source_file": excel_path.name,
                "source_sheet": sheet.title,
                "source_row": source_row,
            }
        )

    id_rows: dict[int, list[int]] = defaultdict(list)
    name_rows: dict[str, list[dict[str, int]]] = defaultdict(list)
    for row in rows:
        id_rows[int(row["case_id"])].append(int(row["source_row"]))
        if row["operator_name"]:
            name_rows[str(row["operator_name"])].append(
                {"case_id": int(row["case_id"]), "source_row": int(row["source_row"])}
            )

    duplicate_ids = {str(key): value for key, value in id_rows.items() if len(value) > 1}
    duplicate_names = {key: value for key, value in name_rows.items() if len(value) > 1}
    source_counts = Counter(str(row["source_status"]) for row in rows)
    executable_count = sum(
        1 for row in rows if row["enabled"] == "true" and row["source_status"] == "READY"
    )
    missing_code_count = sum(1 for row in rows if not str(row["code"]).strip())
    missing_operator_count = sum(1 for row in rows if not str(row["operator_name"]).strip())

    report = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_file": excel_path.name,
        "source_sheet": sheet.title,
        "total_records": len(rows),
        "executable_records": executable_count,
        "missing_code_records": missing_code_count,
        "missing_operator_name_records": missing_operator_count,
        "source_status_counts": dict(sorted(source_counts.items())),
        "duplicate_case_ids": duplicate_ids,
        "duplicate_operator_names": duplicate_names,
        "syntax_errors": syntax_errors,
        "warnings": [],
        "errors": [],
    }
    if duplicate_names:
        report["warnings"].append("发现重复英文算子名；保留全部 case_id，并在运行时按 case_id 独立追踪。")
    if duplicate_ids:
        report["errors"].append("case_id 必须唯一。")
    if syntax_errors:
        report["errors"].append("部分非空示例代码未通过 Python 语法检查。")
    if not rows:
        report["errors"].append("Excel 未读取到业务记录。")
    report["passed"] = not report["errors"]
    return rows, report


def write_csv(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从 OGE 算子总表生成唯一规范运行输入 operators.csv")
    parser.add_argument("--excel", required=True, help="原始算子 Excel 路径")
    parser.add_argument("--output", default="input/operators.csv", help="输出 CSV 路径")
    parser.add_argument("--sheet", default=None, help="工作表名；默认第一个工作表")
    parser.add_argument("--report", default=None, help="质检 JSON 路径；默认与 CSV 同目录")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    excel_path = Path(args.excel).resolve()
    output_path = Path(args.output).resolve()
    report_path = Path(args.report).resolve() if args.report else output_path.with_suffix(".build_report.json")
    if not excel_path.exists():
        raise SystemExit(f"原始 Excel 不存在: {excel_path}")

    rows, report = build_rows(excel_path, args.sheet)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    if not report["passed"]:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 2

    write_csv(rows, output_path)
    print(f"已生成: {output_path}")
    print(
        "质检统计: "
        f"总数={report['total_records']}，可执行={report['executable_records']}，"
        f"缺代码={report['missing_code_records']}，缺英文名={report['missing_operator_name_records']}"
    )
    if report["duplicate_operator_names"]:
        print("[WARN] 重复英文名: " + ", ".join(report["duplicate_operator_names"].keys()))
    print(f"质检报告: {report_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
